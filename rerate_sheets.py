#!/usr/bin/env python
"""Re-rate every match in Elo Mafia Rankings.xlsx from scratch under a chosen
TrueSkill config and write the results back into the workbook as NEW sheets,
leaving the originals untouched.

Two configs are available, matching the two production backends in
backend/main.py:
  - "main": original mafia site   (ghosts 25.7 / 23.85, beta 5.5)
  - "ego" : ego-mafia adjusted    (ghosts 25.275 / 24.275, beta 5.0)

For each config we recompute, starting every player fresh at mu=25,
sigma=25/3 and replaying the current season's games (GameID 46-162) in
chronological order, exactly as backend/main.py would:
  - <Config> MatchHistory : per-player rows, newest game first, with
    recomputed RateChange / old_mu / new_mu / new_sigma / old_rating /
    new_rating / old_sigma
  - <Config> Stats Summary : per-player aggregates with recomputed final
    mu / sigma / Rating (win/loss counts are historical and unchanged)

The win/loss outcomes are historical facts and never change — only the
ratings derived from them do.

Usage:
    python rerate_sheets.py            # writes both ego + main sheets, validates
    python rerate_sheets.py --validate # validate main re-rate vs original only
"""

import argparse
import math
from dataclasses import dataclass

import pandas as pd
from openpyxl import load_workbook
from trueskill import Rating, TrueSkill

XLSX = "Elo Mafia Rankings.xlsx"
SOURCE_HISTORY = "MatchHistory"
SOURCE_STATS = "Stats Summary"

PRIOR_MU = 25.0
PRIOR_SIGMA = 25.0 / 3.0

HISTORY_COLS = [
    "GameID", "Player", "Alignment", "Result", "RateChange",
    "old_mu", "new_mu", "new_sigma", "old_rating", "new_rating", "old_sigma",
]


@dataclass(frozen=True)
class Config:
    key: str
    label: str
    mafia_ghost_mu: float
    town_ghost_mu: float
    beta: float
    ghost_sigma: float = 0.8

    @property
    def env(self):
        return TrueSkill(tau=0.1, beta=self.beta, draw_probability=0.0)


CONFIGS = {
    "main": Config("main", "Main (25.7/23.85 b5.5)", 25.7, 23.85, 5.5),
    "ego": Config("ego", "Ego (25.275/24.275 b5.0)", 25.275, 24.275, 5.0),
}

# Games where a player was logged on BOTH teams (data-entry error). Dan
# confirmed the true alignment; the other row is dropped. Other duplicate
# games (136, 151, 155) are exact-duplicate rows resolved by dedup-keep-first.
DUP_KEEP_ALIGNMENT = {(83, "Whyin"): "Mafia", (116, "Ken"): "Mafia"}


def display_rating(mu, sigma):
    return round((mu - 1.5 * sigma) * 68)


def compute_trueskill(mafia, town, mafia_won, cfg):
    """Ghost-padded TrueSkill update. `mafia`/`town` are lists of
    {name, mu, sigma}. Returns {name: {mu, sigma}} for real players only."""
    env = cfg.env
    n_m, n_t = len(mafia), len(town)

    mafia_dict = {}
    mu_geo = sigma_geo = 1.0
    for p in mafia:
        mafia_dict[p["name"]] = Rating(p["mu"], p["sigma"])
        mu_geo *= p["mu"]
        sigma_geo *= p["sigma"]
    mu_avg = mu_geo ** (1 / n_m)
    sigma_avg = sigma_geo ** (1 / n_m)
    for i in range(n_t - n_m):
        mafia_dict[f"_mafia_avg{i}"] = Rating(mu_avg, sigma_avg)
    for i in range(n_t):
        mafia_dict[f"_mafia_ghost{i}"] = Rating(cfg.mafia_ghost_mu, cfg.ghost_sigma)

    town_dict = {}
    for p in town:
        town_dict[p["name"]] = Rating(p["mu"], p["sigma"])
    for i in range(n_t):
        town_dict[f"_town_ghost{i}"] = Rating(cfg.town_ghost_mu, cfg.ghost_sigma)

    ranks = [0, 1] if mafia_won else [1, 0]
    rated = env.rate([mafia_dict, town_dict], ranks=ranks)
    out = {}
    for p in mafia:
        r = rated[0][p["name"]]
        out[p["name"]] = {"mu": r.mu, "sigma": r.sigma}
    for p in town:
        r = rated[1][p["name"]]
        out[p["name"]] = {"mu": r.mu, "sigma": r.sigma}
    return out


def rerate(mh, cfg):
    """Replay all games chronologically under cfg.

    Returns (history_rows oldest-first, final ratings dict)."""
    ratings = {}
    history_rows = []
    for gid in sorted(mh["GameID"].unique()):
        g = mh[mh["GameID"] == gid]
        # Drop the spurious team row for Dan-confirmed both-teams mis-logs,
        # then dedupe exact-duplicate rows by keeping the first occurrence.
        for (dgid, dplayer), keep_align in DUP_KEEP_ALIGNMENT.items():
            if dgid == gid:
                g = g[~((g["Player"] == dplayer) & (g["Alignment"] != keep_align))]
        g = g.drop_duplicates(subset=["Player"], keep="first")
        rated = g[g["Result"].isin(["Win", "Loss"])]
        mafia_won = (rated.loc[rated["Alignment"] == "Mafia", "Result"] == "Win").any()

        mafia, town = [], []
        for _, r in rated.iterrows():
            cr = ratings.get(r["Player"], {"mu": PRIOR_MU, "sigma": PRIOR_SIGMA})
            entry = {"name": r["Player"], "mu": cr["mu"], "sigma": cr["sigma"]}
            (mafia if r["Alignment"] == "Mafia" else town).append(entry)

        new = compute_trueskill(mafia, town, mafia_won, cfg)

        # Emit rows in the sheet's original within-game order
        for _, r in g.iterrows():
            name, align, result = r["Player"], r["Alignment"], r["Result"]
            if result not in ("Win", "Loss"):
                history_rows.append([gid, name, align, result, 0, "", "", "", "", "", ""])
                continue
            old = ratings.get(name, {"mu": PRIOR_MU, "sigma": PRIOR_SIGMA})
            nr = new[name]
            old_rtg = display_rating(old["mu"], old["sigma"])
            new_rtg = display_rating(nr["mu"], nr["sigma"])
            history_rows.append([
                gid, name, align, result, new_rtg - old_rtg,
                old["mu"], nr["mu"], nr["sigma"], old_rtg, new_rtg, old["sigma"],
            ])
            ratings[name] = {"mu": nr["mu"], "sigma": nr["sigma"]}
    return history_rows, ratings


def build_stats(mh, ratings):
    """Per-player aggregates (win/loss counts historical) + re-rated mu/sigma."""
    rows = []
    for name, grp in mh.groupby("Player"):
        rated = grp[grp["Result"].isin(["Win", "Loss"])]
        town = rated[rated["Alignment"] == "Town"]
        mafia = rated[rated["Alignment"] == "Mafia"]
        tg, tw = len(town), int((town["Result"] == "Win").sum())
        mg, mw = len(mafia), int((mafia["Result"] == "Win").sum())
        tot, totw = tg + mg, tw + mw
        r = ratings.get(name, {"mu": PRIOR_MU, "sigma": PRIOR_SIGMA})
        rows.append({
            "Name": name,
            "Town Games": tg, "Town Wins": tw,
            "Town %": round(100 * tg / tot, 1) if tot else 0,
            "Town Win %": round(100 * tw / tg, 1) if tg else 0,
            "Mafia Games": mg, "Mafia Wins": mw,
            "Mafia %": round(100 * mg / tot, 1) if tot else 0,
            "Mafia Win %": round(100 * mw / mg, 1) if mg else 0,
            "Total Games": tot,
            "Total Win %": round(100 * totw / tot, 1) if tot else 0,
            "Mu": r["mu"], "Sigma": r["sigma"],
            "Rating": display_rating(r["mu"], r["sigma"]),
        })
    df = pd.DataFrame(rows).sort_values("Rating", ascending=False).reset_index(drop=True)
    return df


def validate_main(mh):
    """Re-rate under main config, compare to the intact MatchRatings sheet
    (Stats Summary is formula-driven and loses its cache on openpyxl save)."""
    _, ratings = rerate(mh, CONFIGS["main"])
    mr = pd.read_excel(XLSX, sheet_name="MatchRatings")
    mr = mr.rename(columns={mr.columns[0]: "Name", mr.columns[1]: "mu",
                            mr.columns[2]: "sigma"})
    print(f"{'Player':<16}{'origRtg':>8}{'reRtg':>8}{'Δ':>6}"
          f"{'origMu':>9}{'reMu':>9}{'origSig':>9}{'reSig':>9}")
    print("-" * 76)
    max_rtg_diff = 0.0
    n_mismatch = 0
    n = 0
    for _, row in mr.iterrows():
        name = row["Name"]
        if pd.isna(name) or not str(name).strip():
            continue
        n += 1
        orig_rtg = display_rating(row["mu"], row["sigma"])
        if name not in ratings:
            print(f"{name:<16}{orig_rtg:>8}{'(no re-rate — name drift?)':>40}")
            n_mismatch += 1
            continue
        re = ratings[name]
        re_rtg = display_rating(re["mu"], re["sigma"])
        d = re_rtg - orig_rtg
        max_rtg_diff = max(max_rtg_diff, abs(d))
        if abs(d) > 1 or abs(re["mu"] - row["mu"]) > 0.01:
            n_mismatch += 1
            print(f"{name:<16}{orig_rtg:>8}{re_rtg:>8}{d:>+6}"
                  f"{row['mu']:>9.3f}{re['mu']:>9.3f}"
                  f"{row['sigma']:>9.3f}{re['sigma']:>9.3f}")
    print("-" * 76)
    print(f"max rating diff = {max_rtg_diff:.0f},  rows differing >1 rtg "
          f"or >0.01 mu = {n_mismatch}/{n}")
    print("(diffs vs original are expected: history was built under a drifting "
          "live config, not a single static one)")


def write_sheets(mh, cfg):
    hist_rows, ratings = rerate(mh, cfg)
    stats = build_stats(mh, ratings)

    wb = load_workbook(XLSX)
    hist_name = f"{cfg.key.capitalize()} MatchHistory"
    stats_name = f"{cfg.key.capitalize()} Stats Summary"
    for nm in (hist_name, stats_name):
        if nm in wb.sheetnames:
            del wb[nm]

    ws_h = wb.create_sheet(hist_name)
    ws_h.append(HISTORY_COLS)
    for row in reversed(hist_rows):  # newest game first, matching source
        ws_h.append(row)

    ws_s = wb.create_sheet(stats_name)
    ws_s.append(list(stats.columns))
    for _, r in stats.iterrows():
        ws_s.append(list(r))

    wb.save(XLSX)
    print(f"  wrote '{hist_name}' ({len(hist_rows)} rows) and "
          f"'{stats_name}' ({len(stats)} players)")
    top = stats.head(5)[["Name", "Rating", "Mu", "Sigma"]]
    print(top.to_string(index=False))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--validate", action="store_true",
                    help="only validate main re-rate vs original, write nothing")
    ap.add_argument("--configs", nargs="+", default=["ego", "main"],
                    choices=list(CONFIGS))
    args = ap.parse_args()

    mh = pd.read_excel(XLSX, sheet_name=SOURCE_HISTORY)
    n_games = mh["GameID"].nunique()
    print(f"Source: {SOURCE_HISTORY} — {n_games} games, "
          f"GameID {mh.GameID.min()}–{mh.GameID.max()}, {len(mh)} player-rows\n")

    print("VALIDATION: re-rate under MAIN config vs original Stats Summary")
    validate_main(mh)

    if args.validate:
        return

    print()
    for key in args.configs:
        cfg = CONFIGS[key]
        print(f"Writing sheets for {cfg.label}:")
        write_sheets(mh, cfg)
        print()


if __name__ == "__main__":
    main()
