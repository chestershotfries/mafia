#!/usr/bin/env python3
"""Fix misspelled player names in MatchHistory and recalculate ratings.

Reads game data from the Excel export, applies name corrections, re-simulates
all games with the backend TrueSkill logic, and writes corrected data back to
Google Sheets.

Usage:
    python fix_names.py          # dry-run (print changes only)
    python fix_names.py --apply  # apply changes to Google Sheets
"""

import json
import os
import sys
from collections import defaultdict

import gspread
import openpyxl
from google.oauth2.service_account import Credentials
from trueskill import TrueSkill, Rating

# --- TrueSkill config (matches backend/main.py) ---

TRUESKILL_MU = 25.0
TRUESKILL_SIGMA = 25 / 3
MAFIA_GHOST_MU = 25.7
MAFIA_GHOST_SIGMA = 0.8
TOWN_GHOST_MU = 23.85
TOWN_GHOST_SIGMA = 0.8

env = TrueSkill(tau=0.1, beta=5.5, draw_probability=0.00)
env.make_as_global()

SHEET_ID = "1vTc6XAa4beDM4n1syQ22Hs10JGVT9PuHNSoTmY051CQ"
SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]
XLSX_PATH = "/home/dan/mafia/Copy of Elo Mafia Rankings.xlsx"

# --- Name fixes: game_id -> {old_name: new_name} ---

NAME_FIXES = {
    119: {"Smitball": "Smit"},
    120: {"Smittball": "Smit", "Brown Kow": "Brownkow"},
    121: {"BP": "Badpants"},
    124: {"Decode": "Matt (Decode)"},
    126: {"Smitball": "Smit", "Brown Kow": "Brownkow"},
    131: {"Smitball": "Smit"},
    132: {"Smitball": "Smit"},
    133: {"Brown Kow": "Brownkow"},
    134: {"Brown Kow": "Brownkow"},
    136: {"BP": "Badpants"},
    138: {"Striker": "Strik3r", "Brown Kow": "Brownkow"},
    139: {"Striker": "Strik3r", "Brown Kow": "Brownkow"},
    141: {"Brown Kow": "Brownkow", "BP": "Badpants"},
    142: {"Brown Kow": "Brownkow"},
    145: {"BP": "Badpants"},
    146: {"Brown Kow": "Brownkow", "BP": "Badpants"},
    148: {"Brown Kow": "Brownkow", "BP": "Badpants", "schmittball": "Smit"},
    149: {"Brown Kow": "Brownkow", "BP": "Badpants", "AA": "Doublea"},
}

# First affected game — recalculate from here onward
FIRST_AFFECTED_GAME = 119

# Orphaned names to remove from MatchRatings and Stats Summary
ORPHANED_NAMES = ["Smitball", "Decode", "Brown Kow", "BP", "AA", "schmittball", "Striker"]


def display_rating(mu, sigma):
    return round((mu - 1.5 * sigma) * 68)


def compute_backend(mafia_players, town_players, mafia_won):
    """Backend TrueSkill logic (from simulate_all_games.py / main.py)."""
    n_mafia = len(mafia_players)
    n_town = len(town_players)

    mafia_dict = {}
    mu_geo = 1.0
    sigma_geo = 1.0
    for p in mafia_players:
        mafia_dict[p["name"]] = Rating(p["mu"], p["sigma"])
        mu_geo *= p["mu"]
        sigma_geo *= p["sigma"]

    mu_avg = mu_geo ** (1 / n_mafia)
    sigma_avg = sigma_geo ** (1 / n_mafia)
    for i in range(n_town - n_mafia):
        mafia_dict[f"_mafia_avg{i}"] = Rating(mu_avg, sigma_avg)
    for i in range(n_town):
        mafia_dict[f"_mafia_ghost{i}"] = Rating(MAFIA_GHOST_MU, MAFIA_GHOST_SIGMA)

    town_dict = {}
    for p in town_players:
        town_dict[p["name"]] = Rating(p["mu"], p["sigma"])
    for i in range(n_town):
        town_dict[f"_town_ghost{i}"] = Rating(TOWN_GHOST_MU, TOWN_GHOST_SIGMA)

    ranks = [0, 1] if mafia_won else [1, 0]
    rated = env.rate([mafia_dict, town_dict], ranks=ranks)

    result = {}
    for p in mafia_players:
        r = rated[0][p["name"]]
        result[p["name"]] = {"mu": r.mu, "sigma": r.sigma}
    for p in town_players:
        r = rated[1][p["name"]]
        result[p["name"]] = {"mu": r.mu, "sigma": r.sigma}
    return result


def load_games(xlsx_path):
    """Load all games from Excel MatchHistory, applying name fixes."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["MatchHistory"]

    games = defaultdict(list)
    for row in range(2, ws.max_row + 1):
        gid = ws.cell(row, 1).value
        if not gid:
            continue
        gid = int(gid)
        player = ws.cell(row, 2).value
        alignment = ws.cell(row, 3).value
        result = ws.cell(row, 4).value

        # Apply name fixes
        if gid in NAME_FIXES and player in NAME_FIXES[gid]:
            old_name = player
            player = NAME_FIXES[gid][old_name]

        games[gid].append({
            "player": player,
            "alignment": alignment,
            "result": result,
        })

    return dict(games)


def simulate_all(games):
    """Simulate all games chronologically, return per-game rating snapshots.

    Returns:
        ratings: final {name: (mu, sigma)} state
        game_snapshots: {game_id: [{player, alignment, result, old_mu, old_sigma,
                                     new_mu, new_sigma, rate_change, old_rating,
                                     new_rating}, ...]}
    """
    game_ids = sorted(games.keys())
    ratings = {}
    game_snapshots = {}

    for gid in game_ids:
        rows = games[gid]

        # Determine winner
        mafia_result = None
        for r in rows:
            if r["alignment"] == "Mafia" and r["result"] in ("Win", "Loss"):
                mafia_result = r["result"]
                break
        if mafia_result is None:
            continue
        mafia_won = mafia_result == "Win"

        # Build rated player lists (exclude Ghost and Night Zero)
        mafia_p = []
        town_p = []
        for r in rows:
            if r["result"] in ("Ghost", "Night Zero"):
                continue
            name = r["player"]
            mu, sigma = ratings.get(name, (TRUESKILL_MU, TRUESKILL_SIGMA))
            entry = {"name": name, "mu": mu, "sigma": sigma}
            if r["alignment"] == "Mafia":
                mafia_p.append(entry)
            else:
                town_p.append(entry)

        if not mafia_p or not town_p:
            continue

        # Compute new ratings
        new_ratings = compute_backend(mafia_p, town_p, mafia_won)

        # Build snapshot for this game
        snapshot = []
        for r in rows:
            name = r["player"]
            if r["result"] in ("Ghost", "Night Zero"):
                snapshot.append({
                    "player": name,
                    "alignment": r["alignment"],
                    "result": r["result"],
                    "old_mu": None,
                    "old_sigma": None,
                    "new_mu": None,
                    "new_sigma": None,
                    "rate_change": None,
                    "old_rating": None,
                    "new_rating": None,
                })
            else:
                old_mu, old_sigma = ratings.get(name, (TRUESKILL_MU, TRUESKILL_SIGMA))
                new_mu = new_ratings[name]["mu"]
                new_sigma = new_ratings[name]["sigma"]
                old_rating = display_rating(old_mu, old_sigma)
                new_rating = display_rating(new_mu, new_sigma)
                snapshot.append({
                    "player": name,
                    "alignment": r["alignment"],
                    "result": r["result"],
                    "old_mu": old_mu,
                    "old_sigma": old_sigma,
                    "new_mu": new_mu,
                    "new_sigma": new_sigma,
                    "rate_change": new_rating - old_rating,
                    "old_rating": old_rating,
                    "new_rating": new_rating,
                })
                ratings[name] = (new_mu, new_sigma)

        game_snapshots[gid] = snapshot

    return ratings, game_snapshots


def get_sheets():
    """Connect to Google Sheets."""
    key_file = os.environ.get("SERVICE_ACCOUNT_KEY_FILE", "service-account.json")
    creds = Credentials.from_service_account_file(key_file, scopes=SCOPES)
    gc = gspread.authorize(creds)
    return gc.open_by_key(SHEET_ID)


def find_sheet_rows(ws_data, game_ids_to_find):
    """Map game_id -> list of (sheet_row_index, player_name) for MatchHistory.

    ws_data is from ws.get_all_values(), row 0 = header.
    Sheet rows are 1-indexed, so data row i corresponds to sheet row i+1.
    """
    game_rows = defaultdict(list)
    for i, row in enumerate(ws_data[1:], start=2):  # sheet row 2 onward
        if not row[0]:
            continue
        gid = int(float(row[0]))
        if gid in game_ids_to_find:
            game_rows[gid].append((i, row[1]))  # (sheet_row, player_name)
    return dict(game_rows)


def main():
    apply = "--apply" in sys.argv

    print("Loading games from Excel...")
    games = load_games(XLSX_PATH)
    game_ids = sorted(games.keys())
    print(f"Loaded {len(game_ids)} games: {game_ids[0]} to {game_ids[-1]}")

    print("\nSimulating all games with corrected names...")
    final_ratings, snapshots = simulate_all(games)

    # Show what changed for affected games
    affected_games = [gid for gid in game_ids if gid >= FIRST_AFFECTED_GAME]
    print(f"\nAffected games (>= {FIRST_AFFECTED_GAME}): {affected_games}")

    # Collect all players who participated in affected games
    affected_players = set()
    for gid in affected_games:
        for entry in snapshots[gid]:
            if entry["new_mu"] is not None:
                affected_players.add(entry["player"])

    print(f"Affected players: {sorted(affected_players)}")

    # Show name fixes
    print("\n--- Name fixes ---")
    for gid, fixes in sorted(NAME_FIXES.items()):
        for old, new in fixes.items():
            print(f"  Game {gid}: {old!r} -> {new!r}")

    # Show rating changes for affected games
    print("\n--- Rating changes for affected games ---")
    for gid in affected_games:
        snap = snapshots[gid]
        print(f"\n  Game {gid}:")
        for entry in snap:
            if entry["new_mu"] is None:
                print(f"    {entry['player']:<20} {entry['result']}")
            else:
                print(f"    {entry['player']:<20} {entry['result']:<8} "
                      f"mu: {entry['old_mu']:.4f} -> {entry['new_mu']:.4f}  "
                      f"rating: {entry['old_rating']} -> {entry['new_rating']} "
                      f"({entry['rate_change']:+d})")

    # Show final ratings for affected players
    print("\n--- Final ratings for affected players ---")
    for name in sorted(affected_players):
        mu, sigma = final_ratings[name]
        rating = display_rating(mu, sigma)
        print(f"  {name:<20} mu={mu:.4f} sigma={sigma:.4f} rating={rating}")

    # Show orphaned entries to delete
    print("\n--- Orphaned entries to delete ---")
    for name in ORPHANED_NAMES:
        if name in final_ratings:
            print(f"  WARNING: {name} still has ratings — should NOT be deleted")
        else:
            print(f"  {name} — will be removed from MatchRatings and Stats Summary")

    if not apply:
        print("\n*** DRY RUN — no changes made. Use --apply to write to Google Sheets. ***")
        return

    # --- Apply to Google Sheets ---
    print("\nConnecting to Google Sheets...")
    ss = get_sheets()

    # --- Update MatchHistory ---
    print("Updating MatchHistory...")
    ws_mh = ss.worksheet("MatchHistory")
    mh_data = ws_mh.get_all_values()

    # Find rows for affected games
    game_row_map = find_sheet_rows(mh_data, set(affected_games))

    batch_updates = []
    for gid in affected_games:
        snap = snapshots[gid]
        sheet_rows = game_row_map.get(gid, [])

        if len(sheet_rows) != len(snap):
            print(f"  WARNING: Game {gid} row count mismatch: "
                  f"sheet={len(sheet_rows)} sim={len(snap)}")
            continue

        for (sheet_row, _old_name), entry in zip(sheet_rows, snap):
            # Column B (2): Player name
            batch_updates.append({
                "range": f"B{sheet_row}",
                "values": [[entry["player"]]],
            })

            if entry["new_mu"] is not None:
                # Columns E-K: RateChange, old_mu, new_mu, new_sigma,
                #               old_rating, new_rating, old_sigma
                batch_updates.append({
                    "range": f"E{sheet_row}:K{sheet_row}",
                    "values": [[
                        entry["rate_change"],
                        entry["old_mu"],
                        entry["new_mu"],
                        entry["new_sigma"],
                        entry["old_rating"],
                        entry["new_rating"],
                        entry["old_sigma"],
                    ]],
                })
            else:
                # Unrated player — clear rating columns
                batch_updates.append({
                    "range": f"E{sheet_row}:K{sheet_row}",
                    "values": [["", "", "", "", "", "", ""]],
                })

    ws_mh.batch_update(batch_updates, value_input_option="RAW")
    print(f"  Updated {len(batch_updates)} cell ranges across {len(affected_games)} games")

    # --- Update MatchRatings ---
    print("Updating MatchRatings...")
    ws_mr = ss.worksheet("MatchRatings")
    mr_data = ws_mr.get_all_values()

    # Build name -> row index map
    mr_name_rows = {}
    for i, row in enumerate(mr_data[1:], start=2):
        if row[0]:
            mr_name_rows[row[0]] = i

    mr_updates = []
    for name in affected_players:
        mu, sigma = final_ratings[name]
        if name in mr_name_rows:
            row_idx = mr_name_rows[name]
            mr_updates.append({
                "range": f"B{row_idx}:C{row_idx}",
                "values": [[mu, sigma]],
            })

    # Delete orphaned entries (process in reverse row order to avoid shifting)
    orphan_rows = []
    for name in ORPHANED_NAMES:
        if name in mr_name_rows:
            orphan_rows.append(mr_name_rows[name])
    for row_idx in sorted(orphan_rows, reverse=True):
        ws_mr.delete_rows(row_idx)
        print(f"  Deleted MatchRatings row {row_idx}")

    # Re-fetch data after deletions since row indices shifted
    if orphan_rows:
        mr_data = ws_mr.get_all_values()
        mr_name_rows = {}
        for i, row in enumerate(mr_data[1:], start=2):
            if row[0]:
                mr_name_rows[row[0]] = i
        mr_updates = []
        for name in affected_players:
            mu, sigma = final_ratings[name]
            if name in mr_name_rows:
                row_idx = mr_name_rows[name]
                mr_updates.append({
                    "range": f"B{row_idx}:C{row_idx}",
                    "values": [[mu, sigma]],
                })

    ws_mr.batch_update(mr_updates, value_input_option="RAW")
    print(f"  Updated ratings for {len(mr_updates)} players")

    # --- Update Stats Summary ---
    # Columns B-L are all formula-driven from MatchHistory, so once
    # MatchHistory is corrected the formulas auto-recalculate.
    # We only need to delete orphaned rows and re-sort.
    print("Updating Stats Summary...")
    ws_ss = ss.worksheet("Stats Summary")
    ss_data = ws_ss.get_all_values()

    # Build name -> row index map
    ss_name_rows = {}
    for i, row in enumerate(ss_data[1:], start=2):
        if row[0]:
            ss_name_rows[row[0]] = i

    # Delete orphaned entries (reverse order)
    orphan_rows_ss = []
    for name in ORPHANED_NAMES:
        if name in ss_name_rows:
            orphan_rows_ss.append(ss_name_rows[name])
    for row_idx in sorted(orphan_rows_ss, reverse=True):
        ws_ss.delete_rows(row_idx)
        print(f"  Deleted Stats Summary row {row_idx}")

    # Sort Stats Summary by rating descending
    last_row = len(ws_ss.get_all_values())
    if last_row > 1:
        ws_ss.sort((12, "des"), range=f"A2:L{last_row}")
    print("  Sorted Stats Summary by rating")

    print("\nDone!")


if __name__ == "__main__":
    main()
