#!/usr/bin/env python3
"""Simulate all games from MatchHistory through both backend and experimental
TrueSkill logic.

Reads the Excel MatchHistory sheet, processes games chronologically,
and compares computed ratings against the spreadsheet's stored values.

Usage: python simulate_all_games.py
"""

import math
import sys
from collections import defaultdict

import openpyxl
from trueskill import TrueSkill, Rating

# --- TrueSkill config (matches backend/main.py) ---

TRUESKILL_MU = 25.0
TRUESKILL_SIGMA = 25 / 3
MAFIA_GHOST_MU = 25.7
MAFIA_GHOST_SIGMA = 0.8
TOWN_GHOST_MU = 23.85
TOWN_GHOST_SIGMA = 0.8

env = TrueSkill(tau=0.1, beta=5.5, draw_probability=0.00)
env.make_as_global()


def display_rating(mu, sigma):
    return round((mu - 1.5 * sigma) * 68)


def compute_backend(mafia_players, town_players, mafia_won):
    """Backend logic (backend/main.py compute_trueskill)."""
    n_mafia = len(mafia_players)
    n_town = len(town_players)

    mafia_dict = {}
    mu_geo = 1.0
    sigma_geo = 1.0
    for p in mafia_players:
        mafia_dict[p['name']] = Rating(p['mu'], p['sigma'])
        mu_geo *= p['mu']
        sigma_geo *= p['sigma']

    mu_avg = mu_geo ** (1 / n_mafia)
    sigma_avg = sigma_geo ** (1 / n_mafia)
    for i in range(n_town - n_mafia):
        mafia_dict[f'_mafia_avg{i}'] = Rating(mu_avg, sigma_avg)
    for i in range(n_town):
        mafia_dict[f'_mafia_ghost{i}'] = Rating(MAFIA_GHOST_MU, MAFIA_GHOST_SIGMA)

    town_dict = {}
    for p in town_players:
        town_dict[p['name']] = Rating(p['mu'], p['sigma'])
    for i in range(n_town):
        town_dict[f'_town_ghost{i}'] = Rating(TOWN_GHOST_MU, TOWN_GHOST_SIGMA)

    ranks = [0, 1] if mafia_won else [1, 0]
    rated = env.rate([mafia_dict, town_dict], ranks=ranks)

    result = {}
    for p in mafia_players:
        r = rated[0][p['name']]
        result[p['name']] = {'mu': r.mu, 'sigma': r.sigma}
    for p in town_players:
        r = rated[1][p['name']]
        result[p['name']] = {'mu': r.mu, 'sigma': r.sigma}
    return result


def compute_experimental(mafia_players, town_players, mafia_won):
    """Experimental script logic (Mafia_Rating_Experimental.py rate_game).

    Key differences from backend:
    - Builds town_dict BEFORE mafia_dict
    - Uses math.pow() for geometric mean
    - Different ghost/filler key names (no underscore prefix)
    - Filler count based on len(mafia_dict) not len(mafia_players)
    """
    mafia_dict = {}
    town_dict = {}

    # Town first (matches experimental script order)
    for p in town_players:
        town_dict[p['name']] = Rating(p['mu'], p['sigma'])

    for i in range(len(town_players)):
        town_dict[f'town_ghost{i}'] = Rating(TOWN_GHOST_MU, TOWN_GHOST_SIGMA)

    # Then mafia
    mafia_mu_geo_total = 1.0
    mafia_sigma_geo_total = 1.0
    for p in mafia_players:
        mafia_mu_geo_total *= p['mu']
        mafia_sigma_geo_total *= p['sigma']
        mafia_dict[p['name']] = Rating(p['mu'], p['sigma'])

    mafia_mu_avg = math.pow(mafia_mu_geo_total, (1 / len(mafia_players)))
    mafia_sigma_avg = math.pow(mafia_sigma_geo_total, (1 / len(mafia_players)))

    # Fillers: len(town_players) - len(mafia_dict) — mafia_dict has real players only here
    for i in range(len(town_players) - len(mafia_dict)):
        mafia_dict[f'mafia_avg{i}'] = Rating(mafia_mu_avg, mafia_sigma_avg)

    for i in range(len(town_players)):
        mafia_dict[f'mafia_ghost{i}'] = Rating(MAFIA_GHOST_MU, MAFIA_GHOST_SIGMA)

    rating_groups = [mafia_dict, town_dict]

    if mafia_won:
        rated = env.rate(rating_groups, ranks=[0, 1])
    else:
        rated = env.rate(rating_groups, ranks=[1, 0])

    result = {}
    for p in mafia_players:
        r = rated[0][p['name']]
        result[p['name']] = {'mu': r.mu, 'sigma': r.sigma}
    for p in town_players:
        r = rated[1][p['name']]
        result[p['name']] = {'mu': r.mu, 'sigma': r.sigma}
    return result


def load_games(xlsx_path):
    """Load all games from MatchHistory, return dict of game_id -> rows."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["MatchHistory"]

    games = defaultdict(list)
    for row in range(2, ws.max_row + 1):
        gid = ws.cell(row, 1).value
        if not gid:
            continue
        gid = int(gid)
        player = ws.cell(row, 2).value
        alignment = ws.cell(row, 3).value
        result = ws.cell(row, 4).value
        old_mu = ws.cell(row, 6).value
        new_mu = ws.cell(row, 7).value
        new_sigma = ws.cell(row, 8).value
        old_sigma = ws.cell(row, 11).value

        games[gid].append({
            'player': player,
            'alignment': alignment,
            'result': result,
            'old_mu': float(old_mu) if old_mu else None,
            'new_mu': float(new_mu) if new_mu else None,
            'new_sigma': float(new_sigma) if new_sigma else None,
            'old_sigma': float(old_sigma) if old_sigma else None,
        })

    return dict(games)


def main():
    xlsx_path = "/home/dan/mafia/Copy of Elo Mafia Rankings.xlsx"
    games = load_games(xlsx_path)
    game_ids = sorted(games.keys())

    print(f"Loaded {len(game_ids)} games: {game_ids[0]} to {game_ids[-1]}")

    # Two independent rating states — both start fresh
    ratings_back = {}   # backend logic
    ratings_exp = {}    # experimental logic

    back_diffs = []
    exp_diffs = []

    for gid in game_ids:
        rows = games[gid]

        # Determine winner from mafia result
        mafia_result = None
        for r in rows:
            if r['alignment'] == 'Mafia' and r['result'] in ('Win', 'Loss'):
                mafia_result = r['result']
                break
        if mafia_result is None:
            print(f"  Game {gid}: SKIP (no mafia win/loss found)")
            continue
        mafia_won = mafia_result == 'Win'

        # Build rated player lists for BOTH methods (exclude Ghost and Night Zero)
        mafia_p_back = []
        town_p_back = []
        mafia_p_exp = []
        town_p_exp = []
        for r in rows:
            if r['result'] in ('Ghost', 'Night Zero'):
                continue
            name = r['player']
            bmu, bsig = ratings_back.get(name, (TRUESKILL_MU, TRUESKILL_SIGMA))
            emu, esig = ratings_exp.get(name, (TRUESKILL_MU, TRUESKILL_SIGMA))
            if r['alignment'] == 'Mafia':
                mafia_p_back.append({'name': name, 'mu': bmu, 'sigma': bsig})
                mafia_p_exp.append({'name': name, 'mu': emu, 'sigma': esig})
            else:
                town_p_back.append({'name': name, 'mu': bmu, 'sigma': bsig})
                town_p_exp.append({'name': name, 'mu': emu, 'sigma': esig})

        if not mafia_p_back or not town_p_back:
            print(f"  Game {gid}: SKIP (empty team)")
            continue

        # Compute both
        nr_back = compute_backend(mafia_p_back, town_p_back, mafia_won)
        nr_exp = compute_experimental(mafia_p_exp, town_p_exp, mafia_won)

        # Compare to spreadsheet and update state
        back_max = 0.0
        exp_max = 0.0
        for r in rows:
            if r['result'] in ('Ghost', 'Night Zero'):
                continue
            name = r['player']
            ratings_back[name] = (nr_back[name]['mu'], nr_back[name]['sigma'])
            ratings_exp[name] = (nr_exp[name]['mu'], nr_exp[name]['sigma'])

            if r['new_mu'] is not None:
                back_max = max(back_max, abs(nr_back[name]['mu'] - r['new_mu']))
                exp_max = max(exp_max, abs(nr_exp[name]['mu'] - r['new_mu']))

        winner = "Mafia" if mafia_won else "Town"
        n_m = len(mafia_p_back)
        n_t = len(town_p_back)

        # Check if backend and experimental differ from each other
        be_max = 0.0
        for r in rows:
            if r['result'] in ('Ghost', 'Night Zero'):
                continue
            name = r['player']
            be_max = max(be_max, abs(nr_back[name]['mu'] - nr_exp[name]['mu']))

        back_diffs.append((gid, back_max, exp_max, be_max, n_m, n_t, winner))

    # Summary
    print(f"\n{'='*80}")
    print(f"RESULTS: backend vs experimental vs spreadsheet")
    print(f"{'='*80}")

    print(f"\n{'Game':>6} {'back-sheet':>12} {'exp-sheet':>12} {'back-exp':>10} {'M':>3} {'T':>3} {'Winner':>8}")
    print('-' * 60)
    for gid, bd, ed, bed, nm, nt, winner in back_diffs:
        print(f"{gid:>6} {bd:>12.8f} {ed:>12.8f} {bed:>10.2e} {nm:>3} {nt:>3} {winner:>8}")

    # Final ratings comparison
    print(f"\n{'='*80}")
    print(f"FINAL RATINGS COMPARISON (after game {game_ids[-1]})")
    print(f"{'='*80}")
    print(f"{'Player':<20} {'back_mu':>12} {'exp_mu':>12} {'back-exp':>10} {'sheet_rating':>12}")

    # Load final sheet ratings for comparison
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws_mr = wb["MatchRatings"]
    sheet_ratings = {}
    for row in range(2, ws_mr.max_row + 1):
        name = ws_mr.cell(row, 1).value
        if not name:
            continue
        try:
            mu = float(ws_mr.cell(row, 2).value)
            sigma = float(ws_mr.cell(row, 3).value)
            sheet_ratings[name] = (mu, sigma)
        except (TypeError, ValueError):
            continue

    all_names = sorted(set(list(ratings_back.keys()) + list(ratings_exp.keys())))
    for name in all_names:
        bmu, bsig = ratings_back.get(name, (0, 0))
        emu, esig = ratings_exp.get(name, (0, 0))
        diff = bmu - emu
        smu, ssig = sheet_ratings.get(name, (0, 0))
        s_rating = display_rating(smu, ssig) if smu else ''
        print(f"{name:<20} {bmu:>12.6f} {emu:>12.6f} {diff:>+10.2e} {s_rating:>12}")


if __name__ == '__main__':
    main()
